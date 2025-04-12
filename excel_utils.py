import pandas as pd
import os
import openpyxl
from openpyxl.styles.borders import Border

def remove_empty_columns(table: list) -> list:
    """
    表データから空の列を削除する関数。

    Args:
        table (list): 2次元リスト形式の表データ。

    Returns:
        list: 空の列が削除された表データ。
    """
    # 列ごとにデータが存在するかを確認
    non_empty_columns = [
        any(row[i] is not None and row[i] != "" for row in table)
        for i in range(len(table[0]))
    ]

    # 空でない列のみを残す
    return [
        [cell for i, cell in enumerate(row) if non_empty_columns[i]]
        for row in table
    ]

# Excelファイルから罫線で囲まれた複数の表を抽出し、Markdown形式で返す関数
def extract_tables_from_excel(file_path: str) -> dict:
    """
    Excelファイルから罫線で囲まれた複数の表を抽出し、Markdown形式で返す関数。
    表として認識されない箇所は地の文として出力します。
    元のExcelの記載順を保持します。

    Args:
        file_path (str): Excelファイルのパス。

    Returns:
        dict: シート名をキー、Markdown形式の表と地の文を値とする辞書。
    """
    try:
        workbook = openpyxl.load_workbook(file_path)
        content = {}

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            markdown_output = []
            current_table = []

            for row in sheet.iter_rows():
                row_data = []
                skip_row = False

                for cell in row:
                    # セルに罫線があり、取り消し線が引かれていない場合のみ値を取得
                    if cell.border and any(
                        side.style is not None for side in [
                            cell.border.top, cell.border.bottom, cell.border.left, cell.border.right
                        ]
                    ):
                        if cell.font and cell.font.strike:  # 取り消し線が引かれている場合
                            skip_row = True  # この行を無視
                            break
                        else:
                            row_data.append(cell.value if cell.value is not None else "")
                    else:
                        row_data.append(None)

                if skip_row:
                    continue  # 取り消し線がある行はスキップ

                if any(row_data):
                    current_table.append(row_data)
                else:
                    if current_table:  # 現在の表が終了した場合
                        current_table = remove_empty_columns(current_table)  # 空の列を削除

                        markdown_table = []
                        for i, table_row in enumerate(current_table):
                            if i == 0:  # 最初の行がタイトル行の場合
                                markdown_table.append("| " + " | ".join([str(value) if value else "" for value in table_row]) + " |")
                                markdown_table.append("|---" * len(table_row) + "|")  # タイトル行とデータ行の間に区切り線を追加
                            else:
                                markdown_table.append("| " + " | ".join([str(value) if value else "" for value in table_row]) + " |")
                        markdown_output.append("\n".join(markdown_table))
                        current_table = []
                    else:
                        # 表ではない地の文として扱う
                        if any(cell.font and cell.font.bold and cell.font.underline for cell in row):
                            markdown_output.append("### " + "".join([str(cell.value) if cell.value else "" for cell in row]))
                        elif any(cell.font and cell.font.bold and not cell.font.underline for cell in row):
                            markdown_output.append("#### " + "".join([str(cell.value) if cell.value else "" for cell in row]))
                        elif any(cell.font and not cell.font.bold and cell.font.underline for cell in row):
                            markdown_output.append("##### " + "".join([str(cell.value) if cell.value else "" for cell in row]))
                        else:
                            markdown_output.append("".join([str(cell.value) if cell.value else "" for cell in row]))

            # 最後の表を追加
            if current_table:
                current_table = remove_empty_columns(current_table)  # 空の列を削除
                markdown_table = []
                for i, table_row in enumerate(current_table):
                    if i == 0:
                        markdown_table.append("| " + " | ".join([str(value) if value else "" for value in table_row]) + " |")
                        markdown_table.append("|---" * len(table_row) + "|")
                    else:
                        markdown_table.append("| " + " | ".join([str(value) if value else "" for value in table_row]) + " |")
                markdown_output.append("\n".join(markdown_table))

            # シートの内容をまとめる
            content[sheet_name] = "\n\n".join(markdown_output)

        return content
    except Exception as e:
        print(f"エラー: Excelファイルの表抽出に失敗しました: {e}")
        return {}

if __name__ == "__main__":
    # サンプルExcelファイルのパス
    sample_file_path = "sample/外部設計書_画面設計（あいうえお）.xlsx"

    # 表を抽出
    tables = extract_tables_from_excel(sample_file_path)

    # 抽出した表をMarkdown形式で出力
    output_dir = "output"
    os.makedirs(output_dir, exist_ok=True)  # 出力ディレクトリを作成

    output_file_path = os.path.join(output_dir, "content.md")

    if tables:
        with open(output_file_path, "w", encoding="utf-8") as f:
            for sheet_name, content in tables.items():
                f.write(f"## シート名: {sheet_name}\n\n")
                f.write(content + "\n\n")
        print(f"表が {output_file_path} に書き出されました。")
    else:
        print("表が見つかりませんでした。")